import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
import matplotlib.pyplot as plt
from io import BytesIO
import seaborn as sns
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

class ProfessionalExcelAnalytics:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
        # Define professional color scheme
        self.colors = {
            'header': '1F4E79',      # Dark Blue
            'subheader': '2F5F8F',   # Medium Blue
            'accent': 'E7E6E6',      # Light Gray
            'positive': '70AD47',     # Green
            'negative': 'C55A5A',     # Red
            'warning': 'FFC000',      # Amber
            'neutral': '7F7F7F'       # Gray
        }
        
    def load_and_clean_data(self, csv_file='Main4 - Main3.csv'):
        """Load and clean the manufacturing data"""
        try:
            # Read the CSV file
            data = pd.read_csv(csv_file)
            
            # Clean numeric columns
            def clean_currency(x):
                try:
                    if str(x).strip() in ['-', '', 'nan', 'NaN']:
                        return 0.0
                    return float(str(x).replace(',', ''))
                except (ValueError, AttributeError):
                    return 0.0
            
            # Apply cleaning to numeric columns
            numeric_columns = ['Qty', 'Rate', 'Value', 'Fwt',
                             'Target  Manpower', 'variation Manpower', 'Actual Manpower',
                             'Target RawMaterial(Cost)', 'variation RawMaterial', 'Actual RawMaterial',
                             'Target Machinepower(Cost)', 'variation Machine power', 'Actual Machine power',
                             'Target Overhead(Cost)or Profit', 'variation overhead ', 'Actual Overhead or profit']
            
            for col in numeric_columns:
                if col in data.columns:
                    data[col] = data[col].apply(clean_currency)
            
            # Convert date
            data['Date'] = pd.to_datetime(data['Date'], errors='coerce', dayfirst=True)
            
            # Calculate additional metrics
            data['Total_Target_Cost'] = (data['Target  Manpower'] + 
                                       data['Target RawMaterial(Cost)'] + 
                                       data['Target Machinepower(Cost)'] + 
                                       data['Target Overhead(Cost)or Profit'])
            
            data['Total_Actual_Cost'] = (data['Actual Manpower'] + 
                                       data['Actual RawMaterial'] + 
                                       data['Actual Machine power'] + 
                                       data['Actual Overhead or profit'])
            
            data['Cost_Variance'] = data['Total_Actual_Cost'] - data['Total_Target_Cost']
            data['Cost_Variance_Pct'] = (data['Cost_Variance'] / data['Total_Target_Cost']) * 100
            data['Profit_Margin'] = ((data['Value'] - data['Total_Actual_Cost']) / data['Value']) * 100
            data['Unit_Profit'] = (data['Value'] - data['Total_Actual_Cost']) / data['Qty']
            
            return data
            
        except Exception as e:
            print(f"Error loading data: {e}")
            return pd.DataFrame()
    
    def create_executive_summary(self, data):
        """Create Executive Summary Sheet"""
        ws = self.wb.create_sheet("Executive Summary")
        
        # Title
        ws['B2'] = "MANUFACTURING BUSINESS INTELLIGENCE DASHBOARD"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.colors['header'])
        ws.merge_cells('B2:J2')
        ws['B2'].alignment = Alignment(horizontal='center')
        
        # Date range
        date_range = f"Analysis Period: {data['Date'].min().strftime('%B %Y')} - {data['Date'].max().strftime('%B %Y')}"
        ws['B3'] = date_range
        ws['B3'].font = Font(name='Calibri', size=12, italic=True)
        ws.merge_cells('B3:J3')
        ws['B3'].alignment = Alignment(horizontal='center')
        
        # Key Performance Indicators
        current_row = 6
        kpis = [
            ("Total Revenue", f"₹{data['Value'].sum():,.0f}", self.colors['positive']),
            ("Total Orders", f"{len(data):,}", self.colors['header']),
            ("Unique Customers", f"{data['Customer'].nunique()}", self.colors['header']),
            ("Unique Products", f"{data['Part description'].nunique()}", self.colors['header']),
            ("Average Order Value", f"₹{data['Value'].mean():,.0f}", self.colors['positive']),
            ("Total Production Volume", f"{data['Qty'].sum():,.0f} units", self.colors['positive']),
            ("Average Profit Margin", f"{data['Profit_Margin'].mean():.1f}%", 
             self.colors['positive'] if data['Profit_Margin'].mean() > 0 else self.colors['negative']),
            ("Cost Variance", f"{data['Cost_Variance_Pct'].mean():.1f}%", 
             self.colors['negative'] if data['Cost_Variance_Pct'].mean() > 0 else self.colors['positive'])
        ]
        
        # Create KPI cards
        col = 2
        for i, (metric, value, color) in enumerate(kpis):
            if i % 4 == 0 and i > 0:
                current_row += 4
                col = 2
            
            # Metric name
            ws.cell(row=current_row, column=col, value=metric)
            ws.cell(row=current_row, column=col).font = Font(name='Calibri', size=10, bold=True)
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color=self.colors['accent'], 
                                                                   end_color=self.colors['accent'], 
                                                                   fill_type='solid')
            
            # Metric value
            ws.cell(row=current_row + 1, column=col, value=value)
            ws.cell(row=current_row + 1, column=col).font = Font(name='Calibri', size=14, bold=True, color=color)
            ws.cell(row=current_row + 1, column=col).fill = PatternFill(start_color='FFFFFF', 
                                                                        end_color='FFFFFF', 
                                                                        fill_type='solid')
            
            # Add borders
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
            for r in range(current_row, current_row + 2):
                ws.cell(row=r, column=col).border = thin_border
                ws.cell(row=r + 1, column=col).border = thin_border
            
            col += 2
        
        # Customer Analysis Summary
        current_row += 6
        ws.cell(row=current_row, column=2, value="TOP CUSTOMERS BY REVENUE").font = Font(name='Calibri', size=12, bold=True, color=self.colors['header'])
        current_row += 1
        
        top_customers = data.groupby('Customer')['Value'].sum().sort_values(ascending=False).head(5)
        for i, (customer, value) in enumerate(top_customers.items()):
            ws.cell(row=current_row + i, column=2, value=f"{i+1}. {customer}")
            ws.cell(row=current_row + i, column=4, value=f"₹{value:,.0f}")
            ws.cell(row=current_row + i, column=5, value=f"{(value/data['Value'].sum()*100):.1f}%")
        
        # Product Analysis Summary
        ws.cell(row=current_row, column=7, value="TOP PRODUCTS BY VOLUME").font = Font(name='Calibri', size=12, bold=True, color=self.colors['header'])
        
        top_products = data.groupby('Part description')['Qty'].sum().sort_values(ascending=False).head(5)
        for i, (product, qty) in enumerate(top_products.items()):
            ws.cell(row=current_row + i, column=7, value=f"{i+1}. {product[:30]}...")
            ws.cell(row=current_row + i, column=9, value=f"{qty:,.0f}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return ws
    
    def create_financial_analysis(self, data):
        """Create Financial Analysis Sheet"""
        ws = self.wb.create_sheet("Financial Analysis")
        
        # Headers
        ws['B2'] = "FINANCIAL PERFORMANCE ANALYSIS"
        ws['B2'].font = Font(name='Calibri', size=16, bold=True, color=self.colors['header'])
        
        # Monthly Revenue Analysis
        current_row = 5
        monthly_data = data.groupby(data['Date'].dt.to_period('M')).agg({
            'Value': 'sum',
            'Qty': 'sum',
            'Total_Target_Cost': 'sum',
            'Total_Actual_Cost': 'sum',
            'Cost_Variance': 'sum'
        }).reset_index()
        monthly_data['Date'] = monthly_data['Date'].astype(str)
        monthly_data['Profit'] = monthly_data['Value'] - monthly_data['Total_Actual_Cost']
        monthly_data['Profit_Margin_Pct'] = (monthly_data['Profit'] / monthly_data['Value']) * 100
        
        # Add monthly data to sheet
        headers = ['Month', 'Revenue', 'Quantity', 'Target Cost', 'Actual Cost', 'Cost Variance', 'Profit', 'Profit Margin %']
        for i, header in enumerate(headers):
            ws.cell(row=current_row, column=i+2, value=header).font = Font(bold=True, color=self.colors['header'])
            ws.cell(row=current_row, column=i+2).fill = PatternFill(start_color=self.colors['accent'], 
                                                                   end_color=self.colors['accent'], 
                                                                   fill_type='solid')
        
        for i, row in monthly_data.iterrows():
            current_row += 1
            ws.cell(row=current_row, column=2, value=row['Date'])
            ws.cell(row=current_row, column=3, value=row['Value'])
            ws.cell(row=current_row, column=4, value=row['Qty'])
            ws.cell(row=current_row, column=5, value=row['Total_Target_Cost'])
            ws.cell(row=current_row, column=6, value=row['Total_Actual_Cost'])
            ws.cell(row=current_row, column=7, value=row['Cost_Variance'])
            ws.cell(row=current_row, column=8, value=row['Profit'])
            ws.cell(row=current_row, column=9, value=row['Profit_Margin_Pct'])
        
        # Apply conditional formatting for profit margins
        profit_margin_range = f"I6:I{current_row}"
        ws.conditional_formatting.add(profit_margin_range,
                                    CellIsRule(operator='greaterThan', formula=['0'], 
                                              fill=PatternFill(start_color=self.colors['positive'], 
                                                              end_color=self.colors['positive'], 
                                                              fill_type='solid')))
        ws.conditional_formatting.add(profit_margin_range,
                                    CellIsRule(operator='lessThan', formula=['0'], 
                                              fill=PatternFill(start_color=self.colors['negative'], 
                                                              end_color=self.colors['negative'], 
                                                              fill_type='solid')))
        
        # Cost Variance Analysis
        current_row += 3
        ws.cell(row=current_row, column=2, value="COST VARIANCE ANALYSIS BY CATEGORY").font = Font(name='Calibri', size=12, bold=True, color=self.colors['header'])
        current_row += 1
        
        variance_analysis = pd.DataFrame({
            'Category': ['Manpower', 'Raw Material', 'Machine Power', 'Overhead'],
            'Target_Cost': [data['Target  Manpower'].sum(), data['Target RawMaterial(Cost)'].sum(), 
                          data['Target Machinepower(Cost)'].sum(), data['Target Overhead(Cost)or Profit'].sum()],
            'Actual_Cost': [data['Actual Manpower'].sum(), data['Actual RawMaterial'].sum(),
                          data['Actual Machine power'].sum(), data['Actual Overhead or profit'].sum()],
        })
        variance_analysis['Variance'] = variance_analysis['Actual_Cost'] - variance_analysis['Target_Cost']
        variance_analysis['Variance_Pct'] = (variance_analysis['Variance'] / variance_analysis['Target_Cost']) * 100
        
        # Add variance analysis data
        headers = ['Category', 'Target Cost', 'Actual Cost', 'Variance', 'Variance %']
        for i, header in enumerate(headers):
            ws.cell(row=current_row, column=i+2, value=header).font = Font(bold=True, color=self.colors['header'])
        
        for i, row in variance_analysis.iterrows():
            current_row += 1
            ws.cell(row=current_row, column=2, value=row['Category'])
            ws.cell(row=current_row, column=3, value=row['Target_Cost'])
            ws.cell(row=current_row, column=4, value=row['Actual_Cost'])
            ws.cell(row=current_row, column=5, value=row['Variance'])
            ws.cell(row=current_row, column=6, value=row['Variance_Pct'])
        
        return ws
    
    def create_customer_analysis(self, data):
        """Create Customer Analysis Sheet"""
        ws = self.wb.create_sheet("Customer Analysis")
        
        ws['B2'] = "CUSTOMER SEGMENTATION & PROFITABILITY ANALYSIS"
        ws['B2'].font = Font(name='Calibri', size=16, bold=True, color=self.colors['header'])
        
        # Customer performance metrics
        customer_analysis = data.groupby('Customer').agg({
            'Value': ['sum', 'mean', 'count'],
            'Qty': ['sum', 'mean'],
            'Profit_Margin': 'mean',
            'Unit_Profit': 'mean',
            'Date': ['min', 'max']
        }).round(2)
        
        # Flatten column names
        customer_analysis.columns = ['_'.join(col).strip() for col in customer_analysis.columns]
        customer_analysis = customer_analysis.reset_index()
        
        # Calculate customer lifetime and frequency
        customer_analysis['Days_Active'] = (customer_analysis['Date_max'] - customer_analysis['Date_min']).dt.days
        customer_analysis['Order_Frequency'] = customer_analysis['Value_count'] / (customer_analysis['Days_Active'] / 30 + 1)
        
        # Customer segmentation based on revenue and frequency
        revenue_median = customer_analysis['Value_sum'].median()
        frequency_median = customer_analysis['Order_Frequency'].median()
        
        def segment_customer(row):
            if row['Value_sum'] > revenue_median and row['Order_Frequency'] > frequency_median:
                return "Champions"
            elif row['Value_sum'] > revenue_median and row['Order_Frequency'] <= frequency_median:
                return "Loyal Customers"
            elif row['Value_sum'] <= revenue_median and row['Order_Frequency'] > frequency_median:
                return "Potential Loyalists"
            else:
                return "At Risk"
        
        customer_analysis['Segment'] = customer_analysis.apply(segment_customer, axis=1)
        
        # Write customer analysis to sheet
        current_row = 5
        headers = ['Customer', 'Total Revenue', 'Avg Order Value', 'Order Count', 'Total Qty', 
                  'Avg Profit Margin %', 'Order Frequency', 'Segment']
        
        for i, header in enumerate(headers):
            ws.cell(row=current_row, column=i+2, value=header).font = Font(bold=True, color=self.colors['header'])
        
        for i, row in customer_analysis.iterrows():
            current_row += 1
            ws.cell(row=current_row, column=2, value=row['Customer'])
            ws.cell(row=current_row, column=3, value=row['Value_sum'])
            ws.cell(row=current_row, column=4, value=row['Value_mean'])
            ws.cell(row=current_row, column=5, value=row['Value_count'])
            ws.cell(row=current_row, column=6, value=row['Qty_sum'])
            ws.cell(row=current_row, column=7, value=row['Profit_Margin_mean'])
            ws.cell(row=current_row, column=8, value=row['Order_Frequency'])
            ws.cell(row=current_row, column=9, value=row['Segment'])
        
        # Customer segment summary
        current_row += 3
        ws.cell(row=current_row, column=2, value="CUSTOMER SEGMENT SUMMARY").font = Font(name='Calibri', size=12, bold=True, color=self.colors['header'])
        current_row += 1
        
        segment_summary = customer_analysis.groupby('Segment').agg({
            'Value_sum': 'sum',
            'Customer': 'count'
        }).reset_index()
        segment_summary['Revenue_Share'] = (segment_summary['Value_sum'] / segment_summary['Value_sum'].sum()) * 100
        
        headers = ['Segment', 'Customer Count', 'Total Revenue', 'Revenue Share %']
        for i, header in enumerate(headers):
            ws.cell(row=current_row, column=i+2, value=header).font = Font(bold=True, color=self.colors['header'])
        
        for i, row in segment_summary.iterrows():
            current_row += 1
            ws.cell(row=current_row, column=2, value=row['Segment'])
            ws.cell(row=current_row, column=3, value=row['Customer'])
            ws.cell(row=current_row, column=4, value=row['Value_sum'])
            ws.cell(row=current_row, column=5, value=row['Revenue_Share'])
        
        return ws
    
    def create_product_analysis(self, data):
        """Create Product Analysis Sheet"""
        ws = self.wb.create_sheet("Product Analysis")
        
        ws['B2'] = "PRODUCT PERFORMANCE & PORTFOLIO ANALYSIS"
        ws['B2'].font = Font(name='Calibri', size=16, bold=True, color=self.colors['header'])
        
        # Product performance metrics
        product_analysis = data.groupby('Part description').agg({
            'Value': ['sum', 'mean'],
            'Qty': ['sum', 'mean'],
            'Rate': 'mean',
            'Profit_Margin': 'mean',
            'Unit_Profit': 'mean',
            'Customer': 'nunique'
        }).round(2)
        
        product_analysis.columns = ['_'.join(col).strip() for col in product_analysis.columns]
        product_analysis = product_analysis.reset_index()
        
        # Product portfolio analysis (BCG Matrix style)
        revenue_median = product_analysis['Value_sum'].median()
        volume_median = product_analysis['Qty_sum'].median()
        
        def categorize_product(row):
            if row['Value_sum'] > revenue_median and row['Qty_sum'] > volume_median:
                return "Star Products"
            elif row['Value_sum'] > revenue_median and row['Qty_sum'] <= volume_median:
                return "Cash Cows"
            elif row['Value_sum'] <= revenue_median and row['Qty_sum'] > volume_median:
                return "Question Marks"
            else:
                return "Dogs"
        
        product_analysis['Category'] = product_analysis.apply(categorize_product, axis=1)
        product_analysis = product_analysis.sort_values('Value_sum', ascending=False)
        
        # Write top products to sheet
        current_row = 5
        headers = ['Product', 'Total Revenue', 'Total Qty', 'Avg Rate', 'Avg Profit Margin %', 
                  'Customers', 'Category']
        
        for i, header in enumerate(headers):
            ws.cell(row=current_row, column=i+2, value=header).font = Font(bold=True, color=self.colors['header'])
        
        # Show top 50 products
        for i, row in product_analysis.head(50).iterrows():
            current_row += 1
            ws.cell(row=current_row, column=2, value=row['Part description'][:40])
            ws.cell(row=current_row, column=3, value=row['Value_sum'])
            ws.cell(row=current_row, column=4, value=row['Qty_sum'])
            ws.cell(row=current_row, column=5, value=row['Rate_mean'])
            ws.cell(row=current_row, column=6, value=row['Profit_Margin_mean'])
            ws.cell(row=current_row, column=7, value=row['Customer_nunique'])
            ws.cell(row=current_row, column=8, value=row['Category'])
        
        # Product category summary
        current_row += 3
        ws.cell(row=current_row, column=2, value="PRODUCT PORTFOLIO SUMMARY").font = Font(name='Calibri', size=12, bold=True, color=self.colors['header'])
        current_row += 1
        
        category_summary = product_analysis.groupby('Category').agg({
            'Value_sum': 'sum',
            'Part description': 'count',
            'Profit_Margin_mean': 'mean'
        }).reset_index()
        
        headers = ['Category', 'Product Count', 'Total Revenue', 'Avg Profit Margin %']
        for i, header in enumerate(headers):
            ws.cell(row=current_row, column=i+2, value=header).font = Font(bold=True, color=self.colors['header'])
        
        for i, row in category_summary.iterrows():
            current_row += 1
            ws.cell(row=current_row, column=2, value=row['Category'])
            ws.cell(row=current_row, column=3, value=row['Part description'])
            ws.cell(row=current_row, column=4, value=row['Value_sum'])
            ws.cell(row=current_row, column=5, value=row['Profit_Margin_mean'])
        
        return ws
    
    def create_operational_efficiency(self, data):
        """Create Operational Efficiency Analysis Sheet"""
        ws = self.wb.create_sheet("Operational Efficiency")
        
        ws['B2'] = "OPERATIONAL EFFICIENCY & COST VARIANCE ANALYSIS"
        ws['B2'].font = Font(name='Calibri', size=16, bold=True, color=self.colors['header'])
        
        # Efficiency metrics by customer
        current_row = 5
        efficiency_analysis = data.groupby('Customer').agg({
            'Cost_Variance_Pct': 'mean',
            'Profit_Margin': 'mean',
            'Value': 'sum',
            'Total_Target_Cost': 'sum',
            'Total_Actual_Cost': 'sum'
        }).round(2)
        
        efficiency_analysis['Efficiency_Score'] = (100 - abs(efficiency_analysis['Cost_Variance_Pct'])) * 0.5 + efficiency_analysis['Profit_Margin'] * 0.5
        efficiency_analysis = efficiency_analysis.sort_values('Efficiency_Score', ascending=False).reset_index()
        
        headers = ['Customer', 'Cost Variance %', 'Profit Margin %', 'Total Revenue', 'Efficiency Score']
        for i, header in enumerate(headers):
            ws.cell(row=current_row, column=i+2, value=header).font = Font(bold=True, color=self.colors['header'])
        
        for i, row in efficiency_analysis.iterrows():
            current_row += 1
            ws.cell(row=current_row, column=2, value=row['Customer'])
            ws.cell(row=current_row, column=3, value=row['Cost_Variance_Pct'])
            ws.cell(row=current_row, column=4, value=row['Profit_Margin'])
            ws.cell(row=current_row, column=5, value=row['Value'])
            ws.cell(row=current_row, column=6, value=row['Efficiency_Score'])
        
        # Detailed variance analysis by cost category
        current_row += 3
        ws.cell(row=current_row, column=2, value="DETAILED COST VARIANCE ANALYSIS").font = Font(name='Calibri', size=12, bold=True, color=self.colors['header'])
        current_row += 1
        
        variance_detail = pd.DataFrame({
            'Cost_Category': ['Manpower', 'Raw Material', 'Machine Power', 'Overhead'],
            'Total_Target': [data['Target  Manpower'].sum(), data['Target RawMaterial(Cost)'].sum(),
                           data['Target Machinepower(Cost)'].sum(), data['Target Overhead(Cost)or Profit'].sum()],
            'Total_Actual': [data['Actual Manpower'].sum(), data['Actual RawMaterial'].sum(),
                           data['Actual Machine power'].sum(), data['Actual Overhead or profit'].sum()],
            'Avg_Variance_Pct': [data['variation Manpower'].mean(), data['variation RawMaterial'].mean(),
                                data['variation Machine power'].mean(), data['variation overhead '].mean()]
        })
        
        variance_detail['Total_Variance'] = variance_detail['Total_Actual'] - variance_detail['Total_Target']
        variance_detail['Variance_Pct'] = (variance_detail['Total_Variance'] / variance_detail['Total_Target']) * 100
        
        headers = ['Cost Category', 'Target Cost', 'Actual Cost', 'Variance Amount', 'Variance %', 'Avg Variance %']
        for i, header in enumerate(headers):
            ws.cell(row=current_row, column=i+2, value=header).font = Font(bold=True, color=self.colors['header'])
        
        for i, row in variance_detail.iterrows():
            current_row += 1
            ws.cell(row=current_row, column=2, value=row['Cost_Category'])
            ws.cell(row=current_row, column=3, value=row['Total_Target'])
            ws.cell(row=current_row, column=4, value=row['Total_Actual'])
            ws.cell(row=current_row, column=5, value=row['Total_Variance'])
            ws.cell(row=current_row, column=6, value=row['Variance_Pct'])
            ws.cell(row=current_row, column=7, value=row['Avg_Variance_Pct'])
        
        return ws
    
    def create_raw_data_sheet(self, data):
        """Create cleaned raw data sheet"""
        ws = self.wb.create_sheet("Raw Data")
        
        # Add data to worksheet
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color=self.colors['header'])
            cell.fill = PatternFill(start_color=self.colors['accent'], end_color=self.colors['accent'], fill_type='solid')
        
        # Create table
        tab = Table(displayName="RawData", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        return ws
    
    def add_charts_to_summary(self, ws, data):
        """Add charts to executive summary sheet"""
        # Customer sales chart
        customer_data = data.groupby('Customer')['Value'].sum().sort_values(ascending=False).head(10)
        
        # Create chart data in sheet
        chart_start_row = 25
        ws.cell(row=chart_start_row, column=2, value="Customer")
        ws.cell(row=chart_start_row, column=3, value="Sales")
        
        for i, (customer, sales) in enumerate(customer_data.items()):
            ws.cell(row=chart_start_row + i + 1, column=2, value=customer)
            ws.cell(row=chart_start_row + i + 1, column=3, value=sales)
        
        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Top 10 Customers by Sales"
        chart.y_axis.title = 'Sales Value (₹)'
        chart.x_axis.title = 'Customer'
        
        data_range = Reference(ws, min_col=3, min_row=chart_start_row, 
                              max_row=chart_start_row + len(customer_data), max_col=3)
        categories = Reference(ws, min_col=2, min_row=chart_start_row + 1, 
                              max_row=chart_start_row + len(customer_data))
        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        
        ws.add_chart(chart, "F25")
        
        return ws
    
    def generate_professional_dashboard(self):
        """Generate the complete professional Excel dashboard"""
        print("Loading and cleaning data...")
        data = self.load_and_clean_data()
        
        if data.empty:
            print("No data loaded. Please check your CSV file.")
            return
        
        print("Creating Executive Summary...")
        exec_ws = self.create_executive_summary(data)
        exec_ws = self.add_charts_to_summary(exec_ws, data)
        
        print("Creating Financial Analysis...")
        self.create_financial_analysis(data)
        
        print("Creating Customer Analysis...")
        self.create_customer_analysis(data)
        
        print("Creating Product Analysis...")
        self.create_product_analysis(data)
        
        print("Creating Operational Efficiency Analysis...")
        self.create_operational_efficiency(data)
        
        print("Adding Raw Data Sheet...")
        self.create_raw_data_sheet(data)
        
        # Set executive summary as active sheet
        self.wb.active = self.wb["Executive Summary"]
        
        # Save the workbook
        filename = f"Professional_BDM_Analytics_Dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.wb.save(filename)
        print(f"Professional Excel dashboard created: {filename}")
        
        return filename

if __name__ == "__main__":
    # Create the professional analytics dashboard
    analytics = ProfessionalExcelAnalytics()
    dashboard_file = analytics.generate_professional_dashboard()
    print(f"✅ Professional Excel Dashboard completed: {dashboard_file}")